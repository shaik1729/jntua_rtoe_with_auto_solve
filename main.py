from selenium import webdriver
from selenium.webdriver.common.by import By
import pdb
import time
import xlsxwriter
import openpyxl
import os
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
import platform
import easygui


def set_noai(driver):
    uname = easygui.enterbox("Enter the username for noapi: ")
    pwd = easygui.enterbox("Enter the password for noapi: ")
    driver.switch_to.window(driver.window_handles[-1])
    username = driver.find_element(By.XPATH, '//*[@id="root"]/div/div/div/div[2]/form/div/div[1]/div/input')
    username.send_keys(uname)
    password = driver.find_element(By.XPATH, '//*[@id="root"]/div/div/div/div[2]/form/div/div[2]/div/div/input')
    password.send_keys(pwd)
    login_button = driver.find_element(By.ID, 'loginbtn')
    login_button.click()

    setup_api = driver.find_element(By.XPATH, '/html/body/div/div/div[2]/div[1]/main/section/div[2]/div[2]/div[2]/div[2]/div[1]/div/div[2]/a')
    setup_api.click()

def get_result(driver, results_data = [], subjects_with_codes = {}):
    name_and_rollnumber_spans = driver.find_elements(By.CLASS_NAME, 'text-green-800')
    if len(name_and_rollnumber_spans) == 0:
        print("No results found")
        response = {
            "status": "error"
        }
        return response
    roll_number = name_and_rollnumber_spans[0].text
    name = name_and_rollnumber_spans[1].text

    student_result = {
        'roll_number': roll_number,
        'name': name,
        'subjects': []
    }
    
    result_table = driver.find_element(By.XPATH, '//*[@id="__next"]/div[3]/div[2]/div/div/div/div/div[3]/div[2]/table')

    table_body = result_table.find_element(By.TAG_NAME, 'tbody')
    table_body_rows = table_body.find_elements(By.TAG_NAME, 'tr')

    result = {}

    for row in table_body_rows:

        table_body_row_header = row.find_element(By.TAG_NAME, 'th')
        subject_code = table_body_row_header.text

        if subject_code not in subjects_with_codes:
            subject_name = row.find_element(By.CLASS_NAME, 'MuiTableCell-alignLeft').text
            subjects_with_codes[subject_code] = subject_name

        print("Subject code: ", subject_code)

        result[subject_code] = []

        table_body_row_data = row.find_elements(By.TAG_NAME, 'td')

        for data in table_body_row_data:
            result[subject_code].append(data.text)
            print(data.text)
    

    student_result['subjects'] = result
    
    print("Results aded to student_result")

    results_data.append(student_result)

    print("Student result added to results_data")
    response = {
        "status": "success"
    }
    return response


def write_to_excel(results_data = [], subjects_with_codes = [], file_name = ""):

    # writing data to a excel file

    # creating a workbook

    workbook = xlsxwriter.Workbook("{}_{}.xlsx".format(file_name, datetime.now().strftime("%d_%m_%Y_%H_%M_%S")))

    # creating a worksheet

    worksheet = workbook.add_worksheet()

    # write headers based on the subjects_with_codes

    row = 0
    col = 0

    worksheet.write(row, col, "Hall Ticket")
    worksheet.write(row, col + 1, "Name")

    for code, name in subjects_with_codes.items():
        worksheet.write(row, col + 2, code)
        worksheet.write(row, col + 3, "Internal")
        worksheet.write(row, col + 4, "External")
        worksheet.write(row, col + 5, "Total Marks")
        worksheet.write(row, col + 6, "Credits")
        worksheet.write(row, col + 7, "Grade")
        col += 6

    # write results_data to the worksheet based on the subjects_with_codes

    row = 1
    col = 0

    for result in results_data:
        worksheet.write(row, col, result['roll_number'])
        worksheet.write(row, col + 1, result['name'])
        for code, name in subjects_with_codes.items():
            worksheet.write(row, col + 2, result['subjects'][code][4])
            worksheet.write(row, col + 3, result['subjects'][code][1])
            worksheet.write(row, col + 4, result['subjects'][code][2])
            worksheet.write(row, col + 5, result['subjects'][code][3])
            worksheet.write(row, col + 6, result['subjects'][code][5])
            worksheet.write(row, col + 7, result['subjects'][code][6])
            col += 6
        row += 1
        col = 0

    workbook.close()


def begin_fetch(driver, list_of_rollnumbers, results_data = [], subjects_with_codes = {}, file_name = ""):
    for roll_number in list_of_rollnumbers:
        try:
            print("****************************************************************************")
            print("Roll number: ", roll_number)
            roll_number_input_field = driver.find_element(By.NAME, 'hallTicketNumber')

            roll_number_input_field.send_keys(roll_number)

            # wait till data-hcaptcha-response gets populated

            iframe = driver.find_element(By.XPATH, '//*[@id="__next"]/div[3]/div[2]/div/div/div/div/div[2]/form/div/div/div[2]/div/iframe')

            is_chaptcha_solved = False

            while not is_chaptcha_solved:
                try:
                    time.sleep(2)
                    value = iframe.get_attribute('data-hcaptcha-response')
                    if len(value) > 0:
                        is_chaptcha_solved = True
                except Exception as e:
                    print("#####################################################################")
                    print("Exception occured: ", e)
                    print("#####################################################################")
                    pass    

            if(is_chaptcha_solved):
                submit_button = driver.find_element(By.XPATH, '//*[@id="__next"]/div[3]/div[2]/div/div/div/div/div[2]/form/div/div/button')
                submit_button.click()
                response = get_result(driver, results_data, subjects_with_codes)
                if response["status"] == "success":
                    search_new_button = driver.find_element(By.XPATH, '//*[@id="__next"]/div[3]/div[2]/div/div/div/div/div[2]/form/div/div/button')
                    search_new_button.click()
                else:
                    driver.refresh();
                print("Current results count : ", len(results_data))
            else:
                print("Captcha not solved")
        except Exception as e:
            write_to_excel(results_data, subjects_with_codes)
            print("#####################################################################")
            print("Exception occured: ", e)
            print("#####################################################################")
            driver.refresh();

    print("++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++\n\n")
    print(subjects_with_codes)
    print("++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++\n\n")
    print(results_data)
    write_to_excel(results_data, subjects_with_codes, file_name)


def main(driver):
    url = easygui.enterbox("Enter the results url: ")
    driver.get(url)

    results_window = driver.window_handles.index(driver.current_window_handle)

    driver.switch_to.window(driver.window_handles[results_window])

    title = driver.find_element(By.XPATH, '//*[@id="__next"]/div[3]/div[2]/div/div/div/div/div[1]')
    print(title.text)

    # read the rollnumbers list from the excel file

    files_in_current_directory = os.listdir("./excel_files")

    for file in files_in_current_directory:
        list_of_rollnumbers = []
        print("Processing File : ", file)
        path = "./excel_files/" + file
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active

        max_row = sheet_obj.max_row

        for i in range(1, max_row + 1):
            cell_obj = sheet_obj.cell(row = i, column = 1)
            list_of_rollnumbers.append(cell_obj.value)

        if len(list_of_rollnumbers) == 0:
            print("No roll numbers found in the excel file")
        else:
            print("Roll numbers found in the excel file: ", list_of_rollnumbers)
            begin_fetch(driver, list_of_rollnumbers, file_name = file.split(".")[0], results_data = [], subjects_with_codes = {})
    else:
        print("********************************************Finished*************************************************")


if __name__ == "__main__":

    op = Options()
    op.add_extension('./noai.crx')
    platform_name = platform.system()
    if platform_name == "Windows":
        driver = webdriver.Chrome(service=Service("./driver/chromedriver.exe"),options=op)
    else:
        driver = webdriver.Chrome(service=Service("./driver/chromedriver"),options=op)
    driver.maximize_window()

    driver.implicitly_wait(5)

    time.sleep(5)
    set_noai(driver)
    main(driver)
